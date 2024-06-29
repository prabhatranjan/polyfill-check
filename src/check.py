import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup

def check_polyfill(url):
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            scripts = soup.find_all('script')
            for script in scripts:
                src = script.get('src', '')
                if 'polyfill.io' in src:
                    return 'Yes'
                if 'googletagmanager.com' in src:
                    # Check for polyfill.io inside GTM
                    gtm_response = requests.get(src, timeout=10)
                    if 'polyfill.io' in gtm_response.text:
                        return 'Yes'
            return 'No'
        return 'No'
    except requests.RequestException as e:
        print(f"Error checking URL {url}: {e}")
        return 'Error'

# Load the Excel file
file_path = '../files/websites.xlsx'  # Update this with your file path
wb = load_workbook(file_path)
sheet = wb.active

# Iterate over the URLs in the first column
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
    url = row[0]
    result = check_polyfill(url)
    sheet[f'B{row_idx}'] = result

# Save the updated Excel file
wb.save(file_path)
